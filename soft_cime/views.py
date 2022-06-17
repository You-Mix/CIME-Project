from ast import Try
from datetime import date, datetime
from email import message
from heapq import merge
from itertools import product
from multiprocessing import context
from re import S
from tkinter.ttk import Style
from unicodedata import name
from xmlrpc.client import DateTime
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from .models import *
from soft_cime.utils import render_to_pdf
from django.views.generic import View
import matplotlib.pyplot as plt
import urllib, base64
import numpy as np
import io
# from .form import *

from datetime import date

from openpyxl import *

import xlrd
import os
from tempfile import NamedTemporaryFile
from django.db.models import Q
from django.db.models import Sum
# Create your views here.
today = date.today()
mois ={
        1: "Janvier", 
        2: "Fevrier",
        3: "Mars",
        4: "Avril",
        5: "Mai",
        6: "Juin", 
        7: "Juillet",
        8: "Août",
        9: "Septembre",
        10: "Octobre",
        11: "Novembre",
        12: "Décembre",
    }

def connexion(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            login(request, user)
            messages.success(request, "Login succesfully")
            return redirect('index')
        else:
            messages.error(request, "ERREUR !!!")
    return render(request, "soft_cime/login.html")


@login_required
def deconnection(request):
    logout(request)
    return redirect("login")


# @login_required
# def deleteContrib(request, idContrib):
#     contrib = get_object_or_404(contribuable, id=idContrib)
#     if request.method == "POST":
#         contrib.delete()
#         return redirect('liste_contribuable')

#     return render(request, "soft_cime/deleteConfirm.html", {'contribuable':contrib})

@login_required
def index(request):
    if request.user.is_authenticated:
        nbcontrib = contribuable.objects.count()
        personnel = Personnel.objects.get(user=request.user)
        nbDec = Declaration.objects.filter(Q(date_emission__month=today.month) and Q(date_emission__year=today.year)).count()
        nbPaye =Payement.objects.filter(Q(date__month=today.month) and Q(date__year=today.year)).count()
        nbAmr = AMR.objects.filter(Q(date__month=today.month) and Q(date__year=today.year)).count()
        
        import random
        r = lambda: random.randint(0,255)
        couleur_ss = []
        # for i in range(0,5):
        #     couleur.append("("+str(r())+","+str(r())+","+str(r())+")")
            
        ugs = UG.objects.all()
        ug_labels = []
        ug_nbs = []
        for ug in ugs:
            ug_labels.append("UG " + str(ug.ug) + " : " + ug.activite)
            ug_nbs.append(ug.contribuable_set.count())
        
        
        sous_secteurs = Sous_secteur.objects.all()
        ss_labels = []
        ss_nbs = []
        for ss in sous_secteurs:
            ss_labels.append(ss.libelle)
            ss_nbs.append(ss.contribuable_set.count())
            couleur_ss.append("rgb("+str(r())+","+str(r())+","+str(r())+")")
            
        nbdec_mois = []
        labels_month = []
        for i in range(1, 13):
            nbdec_mois.append(Declaration.objects.filter(date_emission__month=i).count())
            labels_month.append(mois[i])
            
        
        
        
        context = {
            "nbcontrib" : nbcontrib,
            "personnel" : personnel,
            "nbDec" : nbDec,
            "nbPaye" : nbPaye,
            "nbdec_mois" : nbdec_mois,
            "nbAmr" : nbAmr,
            "mois" : mois[today.month],
            "today" : today,
            "ug_labels" : ug_labels,
            "ug_nbs" : ug_nbs,
            "labels_month" : labels_month,
            "ss_labels" : ss_labels,
            "ss_nbs" : ss_nbs,
            "couleur_ss" : couleur_ss,
        }
        return render(request, "soft_cime/index.html", context)
    else:
        return redirect('login')





    
# Menu Contribuable
@login_required
def liste_contribuable(request):
    contrib_list = contribuable.objects.all()
    nbContrib = contribuable.objects.count()


    context = {
        'contribuable' : contrib_list,
        'nbContrib' : nbContrib
    }
    
    

    return render(request, "soft_cime/contribuable-list.html", context)


@login_required
def new_contribuable(request):

    # if request.method == 'POST':
    #     fichier = load_workbook(request.POST.get("file"))
    #     sh = fichier["Table1"]
    #     for j in sh:
    #         i=j[0].value
    #         NIU=sh.cell(row=i, column=2).value 
    #         raison_social=sh.cell(row=i, column=3).value 
    #         activite=sh.cell(row=i, column=4).value
    #         regime_impot = sh.cell(row=i, column=5).value
    #         telephone=sh.cell(row=i, column=6).value
    #         arrondissement=sh.cell(row=i, column=7).value
    #         departement=sh.cell(row=i, column=8).value
    #         statut=sh.cell(row=i, column=9).value
    #         ug=sh.cell(row=i, column=10).value
    #         sous_secteur=sh.cell(row=i, column=11).value
                                                     
    #         newContrib = contribuable.objects.create(NIU=NIU, raison_social=raison_social, activite=activite, regime_impot=Regime_impot.objects.get(id=regime_impot),telephone=telephone, arrondissement=arrondissement, departement=Departement.objects.get(id=departement), ug=UG.objects.get(id=ug), sous_secteur=Sous_secteur.objects.get(id=sous_secteur))
    #         newContrib.save()
            

        
    message = " "
    classmsg = ""
    if request.method == 'POST':
        NIU = request.POST.get('NIU')
        raison_social = request.POST.get('raison_social')
        activite = request.POST.get('activite')
        telephone = request.POST.get('telephone')
        arrondissement = request.POST.get('arrondissement')
        departement = Departement.objects.get(departement=request.POST.get('departement'))
        regime_impot = Regime_impot.objects.get(regime_impot=request.POST.get('regime_impot'))
        ug = UG.objects.get(ug=request.POST.get('ug'))
        sous_secteur = Sous_secteur.objects.get(libelle=request.POST.get('sous_secteur'))
        
        if len(contribuable.objects.filter(NIU=NIU))==0:
            newContrib = contribuable.objects.create(NIU=NIU, raison_social=raison_social, activite=activite, regime_impot=regime_impot,telephone=telephone, arrondissement=arrondissement, departement=departement, ug=ug, sous_secteur=sous_secteur)
            if newContrib.save():
                message = "Enregistrer avec succès"
                classmsg = "text-success"
            else:
                message = "Erreur lors de l'enregistrement"
                classmsg = "text-danger"
        else:
            message = "Le contribuable existe dejà"
            classmsg = "text-danger"
        
        
        
        
            
    regimes_impots = Regime_impot.objects.all()
    departements = Departement.objects.all()
    impots = Impot.objects.all()
    ugs = UG.objects.all()
    sous_secteurs = Sous_secteur.objects.all()
    context = {
        'message':message,
        'regimes_impots' : regimes_impots,
        'departements' : departements,
        'impots':impots,
        'ugs':ugs,
        'sous_secteurs':sous_secteurs,
        'classmsg' : classmsg,
    }
    return render(request, "soft_cime/contribuable-new.html", context)


@login_required
def detail_contribuable(request,id):
    
    contrib = contribuable.objects.get(pk=id)
    declarations = Declaration.objects.filter(contribuable=contrib)
    # impots_declare = Impot_Declare.objects.filter(declaration in declarations)
    
    declarations_infos = []
    
    for declaration in declarations:
        line=[]
        line.append(declaration.num_avis)
        line.append(declaration.date_emission)
        montant_dec=Impot_Declare.objects.filter(declaration = declaration).aggregate(Sum('montant'))
        line.append(montant_dec['montant__sum'])
        declarations_infos.append(line)
    
    amrs = AMR.objects.filter(contribuable=contrib)
    # impots_declare = Impot_Declare.objects.filter(declaration in declarations)
    
    amr_infos = []
    
    for amr in amrs:
        line=[]
        line.append(amr.num_amr)
        line.append(amr.date)
        montant_dec=Impot_AMR.objects.filter(amr = amr).aggregate(Sum('montant'))
        line.append(montant_dec['montant__sum'])
        amr_infos.append(line)
        
        
    # total_dec=Impot_Declare.objects.filter(declaration in declarations).aggregate(sum('montant'))
        
    payements = Payement.objects.filter(contribuable=contrib)
    context = {
        'contribuable' : contrib,
        'declarations_infos' : declarations_infos,
        'payements' : payements,
        'declarations' : declarations,
        'amr_infos' : amr_infos,
    }
    
    

    return render(request, "soft_cime/contribuable-detail.html", context)


# Menu Gestion et Suivie
@login_required
def liste_declaration(request):
    declaration_history = Declaration.objects.all().order_by('-date_emission')
    nbdeclaration = Declaration.objects.count()
    impots_declare = Impot_Declare.objects.all()
    


    context = {
        'declarations' : declaration_history,
        'nbdeclaration' : nbdeclaration,
        'impots_declare' : impots_declare,
    }
    
    
    return render(request, "soft_cime/declaration-history.html", context)


@login_required
def declaration_incomplete(request):
    declaration_history = Declaration.objects.filter().order_by('-date_emission')
    nbdeclaration = Declaration.objects.count()
    impots_declare = Impot_Declare.objects.all()
    


    context = {
        'declarations' : declaration_history,
        'nbdeclaration' : nbdeclaration,
        'impots_declare' : impots_declare,
    }
    
    
    return render(request, "soft_cime/declaration-incomplet.html", context)


@login_required
def new_declaration(request):
    #formDec = newDeclaration(request.POST or None)
    contribuables = contribuable.objects.all()
    impots = Impot.objects.all()
    today = date.today()
    message = ""
    classmsg = ""
    
    
    if request.method == 'POST':
        num_avis = request.POST.get("num_avis")
        if len(Declaration.objects.filter(num_avis=num_avis))==0:
            contrib = contribuable.objects.get(NIU=request.POST.get("contribuable"))
            chiffre_affaire = request.POST.get("chiffre_affaire")
            date_limite = request.POST.get("date_limite")
            personnel = Personnel.objects.get(user = request.user)
            
            newDeclaration = Declaration.objects.create(num_avis=num_avis,contribuable=contrib,chiffre_affaire=chiffre_affaire, date_limite=date_limite, personnel=personnel )        
            newDeclaration.save()
            return redirect('new_declaration_impots',newDeclaration.id)
        else:
            message = "Declaration already exist"
            classmsg = "text-danger"
        
    context = {
        'contribuables' : contribuables,
        'impots':impots,
        'today':today,
        'message' : message,
        'classmsg' : classmsg,
    }
    return render(request, "soft_cime/declaration-new.html", context)


@login_required
def new_declaration1(request,idPaye):
    #formDec = newDeclaration(request.POST or None)
    contribuables = contribuable.objects.all()
    impots = Impot.objects.all()
    today = date.today()
    message = ""
    classmsg = ""
    payement = Payement.objects.get(id=idPaye)
    
    if request.method == 'POST':
        num_avis = request.POST.get("num_avis")
        if len(Declaration.objects.filter(num_avis=num_avis))==0:
            contrib = contribuable.objects.get(NIU=request.POST.get("contribuable"))
            chiffre_affaire = request.POST.get("chiffre_affaire")
            date_limite = request.POST.get("date_limite")
            personnel = Personnel.objects.get(user = request.user)
            
            newDeclaration = Declaration.objects.create(num_avis=num_avis,contribuable=contrib,chiffre_affaire=chiffre_affaire, date_limite=date_limite, personnel=personnel )        
            newDeclaration.save()
            return redirect('new_declaration_impots',newDeclaration.id)
        else:
            message = "Declaration already exist"
            classmsg = "text-danger"
        
    context = {
        'contribuables' : contribuables,
        'impots':impots,
        'today':today,
        'message' : message,
        'classmsg' : classmsg,
        'payement' : payement,
        
    }
    return render(request, "soft_cime/declaration-new.html", context)

@login_required
def update_declaration(request, idDec):
    #formDec = newDeclaration(request.POST or None)
    contribuables = contribuable.objects.all()
    declaration = Declaration.objects.get(id=idDec)
    today = date.today()
    
    context = {
        'contribuables' : contribuables,
        'declaration':declaration,
        'today':today,
    }
    
    if request.method == 'POST':
        
        declaration.num_avis = request.POST.get("num_avis")
        declaration.contribuable = contribuable.objects.get(NIU=request.POST.get("contribuable"))
        declaration.chiffre_affaire = request.POST.get("chiffre_affaire")
        declaration.date_limite = request.POST.get("date_limite")
        
        declaration.save()
        
        
        return redirect('new_declaration_impots',declaration.id)
    
    return render(request, "soft_cime/declaration-update.html", context)


@login_required
def new_declaration_impots(request, idDec):
    # formDec = newDeclaration(request.POST or None)
    # impots = contrib.impots.all()
    impots = Impot.objects.all()
    declaration = get_object_or_404(Declaration, pk=idDec)
    impots_declares = Impot_Declare.objects.filter(declaration=declaration).order_by("-id")
    message=""
    classmsg = ""
    if request.method == "POST":
        montant = request.POST.get("montant")
        impot = Impot.objects.get(impot=request.POST.get('impot'))
        exist = Impot_Declare.objects.filter(impot=impot) & Impot_Declare.objects.filter(declaration=declaration)
        if len(exist) == 0:
            newImpotDec = Impot_Declare.objects.create(declaration=declaration, impot=impot, montant=montant)
            newImpotDec.save()
            message = "Enregistré"
            classmsg = "text-success"
        else:
            message = "Impot Déjà Déclaré"
            classmsg = "text-danger"
            
    context = {
        'impots':impots,
        "message":message,
        "classmsg" : classmsg,
        'impots_declares' : impots_declares,
    }
    return render(request, "soft_cime/declaration-new1.html", context)


@login_required
def detail_declaration(request,idDec):
    
    declaration = Declaration.objects.get(id=idDec)
    impots_declares = Impot_Declare.objects.filter(declaration=declaration)
    payement = Payement.objects.filter(num_avis=declaration.num_avis)
    
    statut = "En attente"
    if len(payement)>0:
        for p in payement:
            if p.contribuable == declaration.contribuable:
                statut = "Payé (enregistré le " + str(p.date) + ")"
        
    context = {
        'declaration' : declaration,
        'impots_declares' : impots_declares,
        'statut' : statut,
    }
    
    

    return render(request, "soft_cime/declaration-detail.html", context)

@login_required
def excel_declaration(request,idDec):
    
    declaration = Declaration.objects.get(id=idDec)
    impots_declares = Impot_Declare.objects.filter(declaration=declaration)
    context = {
        'declaration' : declaration,
        'impots_declares' : impots_declares,
    }
    wb=load_workbook("soft_cime/static/doc/detail declaration.xlsx")
    ws = wb.active
    
    ws['B3'] = str(declaration.num_avis)
    ws['B4'] = str(declaration.contribuable.NIU) + " : " + str(declaration.contribuable.raison_social)
    ws['B5'] = str(declaration.chiffre_affaire)
    ws['B6'] = str(declaration.date_emission)
    ws['B7'] = str(declaration.date_limite)
    ws['B8'] = declaration.personnel.nom
    
    paye = Payement.objects.filter(num_avis=declaration.num_avis)
    if len(paye) != 0:
        for p in paye:
            ws['B9'] = "Payé (enregistré le " + str(p.date) + ")"
    else:
        ws['B9'] = "En attente"
    
    # if declaration.num_avis in Payement.objects.all().num_avis:
    #     ws['B9'] = "Payé"
    
    j='A'
    i=13
    for impot in impots_declares:
        ws["A"+str(i)]=impot.impot.impot
        ws["B"+str(i)]=impot.montant
        i = i+1
        # j=chr(ord(j) + 1)

    ws.oddFooter.center.text ="Powered by Le Gestionnaire \n" + str(datetime.today())
    wb.save(os.path.expanduser("~/Downloads/Déclaration " + str(declaration.num_avis) + ".xlsx"))
    os.popen(os.path.expanduser("~/Downloads/Déclaration " + str(declaration.num_avis) + ".xlsx"))
    
    return redirect("detail_declaration", idDec)




# Menu suivi des Recette
@login_required
def liste_payements(request):
    payement_history = Payement.objects.all()
    nbpayement = Payement.objects.count()
    
    declarations = Declaration.objects.all()
    
    nums_avis = []
    for declaration in declarations:
        nums_avis.append(declaration.num_avis)
        
    nb_paye_non = 0
    
    for payement in payement_history:
        if payement.num_avis not in nums_avis:
            nb_paye_non = nb_paye_non + 1
        else:
            declaration = Declaration.objects.get(num_avis = payement.num_avis)
            if payement.contribuable != declaration.contribuable:
                nb_paye_non = nb_paye_non + 1


    context = {
        'payements' : payement_history,
        'nbpayement' : nbpayement,
        'nb_paye_non' : nb_paye_non,
    }
    
    
    return render(request, "soft_cime/payement-history.html", context)

@login_required
def payement_declare(request):
    payement_history = Payement.objects.all()
    declarations = Declaration.objects.all()
    
    paye_declare = []
    n=0
    for payement in payement_history:
        infos = [1, 2]
        for declaration in declarations:
            if (declaration.num_avis == payement.num_avis) and (declaration.contribuable.NIU == payement.contribuable.NIU):
                impots = ""
                for impot in declaration.impots.all():
                    impots = impot.impot + ", "
                    
                infos = [declaration.num_avis, declaration.contribuable.NIU, declaration.date_emission, impots , payement.date, payement.montant]
                paye_declare.append(infos)
                n=n+1
                break


    context = {
        'paye_declare' : paye_declare,
        'nb_payeD' : n,
    }
    
    
    return render(request, "soft_cime/payement-declare.html", context)

@login_required
def payement_non_declare(request):
    
    declarations = Declaration.objects.all()
    
    nums_avis = []
    for declaration in declarations:
        nums_avis.append(declaration.num_avis)
        
    
    payements = Payement.objects.all()
    payement_history=[]
    nb_paye_non = 0
    for payement in payements:
        if payement.num_avis not in nums_avis:
            payement_history.append(payement)
            nb_paye_non = nb_paye_non + 1
        else:
            declaration = Declaration.objects.get(num_avis = payement.num_avis)
            if payement.contribuable != declaration.contribuable:
                payement_history.append(payement)
                nb_paye_non = nb_paye_non + 1
    
    context = {
        'payement_history' : payement_history,
        'nb_paye_non' : nb_paye_non,
    }
    
    
    return render(request, "soft_cime/payement-non_declare.html", context)



@login_required
def new_payement(request):
    contribuables = contribuable.objects.all()
    declarations = Declaration.objects.all()
    payement_history = Payement.objects.all()
    
    
    nums_avis = []
    for declaration in declarations:
        nums_avis.append(declaration.num_avis)
        
    nb_paye_non = 0
    
    for payement in payement_history:
        if payement.num_avis not in nums_avis:
            nb_paye_non = nb_paye_non + 1
        else:
            declaration = Declaration.objects.get(num_avis = payement.num_avis)
            if payement.contribuable != declaration.contribuable:
                nb_paye_non = nb_paye_non + 1
    
    context = {
        'contribuables' : contribuables,
        'declarations' : declarations,
        'today' : today,
        'nb_paye_non' : nb_paye_non,
    }
    
    if request.method == 'POST':
        num_avis = request.POST.get("num_avis")
        contrib = contribuable.objects.get(NIU=request.POST.get("contribuable"))
        montant = request.POST.get("montant")
        date = today
        personnel = Personnel.objects.get(user = request.user)
        date_virement = request.POST.get("date_vir")
        num_virement = request.POST.get("num_vir")
        etat = request.POST.get("etat")
        
        if num_avis == "":
            newPayement = Payement.objects.create(contribuable=contrib,montant=montant, date=date, personnel=personnel, date_virement=date_virement, num_virement=num_virement, etat=etat)        
        else:
            newPayement = Payement.objects.create(num_avis=num_avis,contribuable=contrib,montant=montant, date=date, personnel=personnel, date_virement=date_virement, num_virement=num_virement, etat=etat)        

        newPayement.save()
    
    return render(request, "soft_cime/payement-new.html", context)

# Menu Statistique
@login_required
def stats1(request):
    ugs = UG.objects.all()
    nbContrib = contribuable.objects.count()
    regime_impots = Regime_impot.objects.all()
    departements = Departement.objects.all()
    context = {
        "ugs" : ugs,
        "nbContrib" :nbContrib,
        "regime_impots" : regime_impots,
        "departements" : departements,
    }
    return render(request, "soft_cime/stats1.html", context)

@login_required
def excelStats1(request):
    ugs = UG.objects.all()
    nbContrib = contribuable.objects.count()
    regime_impots = Regime_impot.objects.all()
    departements = Departement.objects.all()
    
    bigTitle = styles.NamedStyle(name = 'bigTitle')
    borderStyle1 = styles.Side(style = 'double', color = '000000')
    bigTitle.border = styles.Border(left = borderStyle1, right = borderStyle1, top = borderStyle1, bottom = borderStyle1)
    bigTitle.font = styles.Font(name = 'Times New Roman', size = 12, bold=True)
    
    headstyle = styles.NamedStyle(name = 'headstyle')
    headstyle.font = styles.Font(name = 'Times New Roman', size = 10, color = 'ffffff', bold=True)
    headstyle.fill = styles.PatternFill(patternType = 'solid', fgColor = '212529')
    borderStyle = styles.Side(style = 'thin', color = 'cccccc')
    headstyle.border = styles.Border(left = borderStyle, right = borderStyle, top = borderStyle, bottom = borderStyle)
    headstyle.alignment.horizontal = "center"
    
    subtitle = styles.NamedStyle(name = 'subtitle')
    subtitle.font = styles.Font(name = 'Times New Roman', size = 10, bold=True)
    subtitle.fill = styles.PatternFill(patternType = 'solid', fgColor = 'cccccc')
    subtitle.border = styles.Border(left = borderStyle, right = borderStyle, top = borderStyle, bottom = borderStyle)
    subtitle.alignment.horizontal = "center"
    
    s2 = styles.NamedStyle(name = 's2')
    s2.alignment.horizontal = "center"
    s2.border = styles.Border(left = borderStyle, right = borderStyle, top = borderStyle, bottom = borderStyle)
    
    s1 = styles.NamedStyle(name = 's1')
    s1.alignment.horizontal = "center"
    s1.font = styles.Font(bold=True)
    s1.border = styles.Border(left = borderStyle, right = borderStyle, top = borderStyle, bottom = borderStyle)
    
    
    
    wb= Workbook()
    ws = wb.active
    
    ws.oddHeader.center.text ="PAR REGIME D'IMPOSITION ET PAR DEPARTEMENT"
    ws.oddHeader.center.font  = "Times New Roman"
    ws.oddHeader.center.size  = 14
    ws.oddHeader.center.border  = borderStyle1
    ws.oddFooter.center.text ="@Le Gestionnaire \n" + str(datetime.today())
    
    ws.print_options.horizontalCentered = True
    
    ws["A1"].border = styles.Border(top = borderStyle1)
    ws.merge_cells("A1:E1")
    
    #PAR SOUS SECTEUR
    ws["A3"] = "SOUS SECTEURS" 
    ws["B3"] = "TAILLES"
    
    ws["A3"].style = headstyle
    ws["B3"].style = headstyle
    
    
    i=4
    for ug in ugs:
        ws.merge_cells("A"+str(i)+":"+"B"+str(i))
        ws["A"+str(i)] = "UG" + str(ug.ug)
        ws["A"+str(i)].style = subtitle
        i=i+1
        for sous_secteur in ug.sous_secteur_set.all():
            ws["A"+str(i)] = sous_secteur.libelle
            ws["A"+str(i)].border = styles.Border(left = borderStyle, right = borderStyle, top = borderStyle, bottom = borderStyle)
            ws["B"+str(i)] = sous_secteur.contribuable_set.count()
            ws["B"+str(i)].style = s2
            i=i+1
        
        ws["A"+str(i)] = "TOTAL " + str(ug.ug)
        ws["A"+str(i)].border = styles.Border(left = borderStyle, right = borderStyle, top = borderStyle, bottom = borderStyle)
        ws["B"+str(i)] = str(ug.contribuable_set.count())
        ws["B"+str(i)].style = s1
        i=i+1
    ws["A"+str(i)] = "TOTAL FICHIER"
    ws["A"+str(i)].style = subtitle
    ws["B"+str(i)] = str(nbContrib)
    ws["B"+str(i)].style = subtitle
    
    ws.column_dimensions['A'].width = 25
    ws.row_dimensions[1].height = 15
    
    #PAR REGIME 
    ws["D3"] = "REGIME D'IMPOSITION" 
    ws["E3"] = "TAILLES"
    
    ws["D3"].style = headstyle
    ws["E3"].style = headstyle
    
    i=4
    for regime in regime_impots:
        ws["D"+str(i)] = regime.regime_impot
        ws["D"+str(i)].border = styles.Border(left = borderStyle, right = borderStyle, top = borderStyle, bottom = borderStyle)
        ws["E"+str(i)] = regime.contribuable_set.count()
        ws["E"+str(i)].style = s1
        i=i+1
        
       
    ws["D"+str(i)] = "TOTAL"
    ws["D"+str(i)].style = subtitle
    ws["E"+str(i)] = str(nbContrib)
    ws["E"+str(i)].style = subtitle
    i=i+2
    
    #PAR DEPARTEMENT
    ws["D"+str(i)] = "DEPARTEMENT" 
    ws["E"+str(i)] = "TAILLES"
    
    ws["D"+str(i)].style = headstyle
    ws["E"+str(i)].style = headstyle
    
    i=i+1
    for departement in departements:
        ws["D"+str(i)] = departement.departement
        ws["D"+str(i)].border = styles.Border(left = borderStyle, right = borderStyle, top = borderStyle, bottom = borderStyle)
        ws["E"+str(i)] = departement.contribuable_set.count()
        ws["E"+str(i)].style = s1
        i=i+1
        
       
    ws["D"+str(i)] = "TOTAL"
    ws["D"+str(i)].style = subtitle
    ws["E"+str(i)] = str(nbContrib)
    ws["E"+str(i)].style = subtitle
    
    
    ws.column_dimensions['D'].width = 25
    
    # with NamedTemporaryFile() as tmp:
    #     wb.save(tmp.name)
    #     tmp.seek(0)
    #     stream = tmp.read()
    
    wb.save(os.path.expanduser("~/Downloads/stats_contribuable.xlsx"))
    os.popen(os.path.expanduser("~/Downloads/stats_contribuable.xlsx"))
    
    return redirect('stats_etats')
    
@login_required    
def stats2(request):
    
    ugs = UG.objects.all()
    nbContrib = contribuable.objects.count()
    
    impots = Impot.objects.all()
    sous_secteurs = Sous_secteur.objects.all()
    
    # if request.method != "POST":
    #     mois=today.month
    # else:
    #     mois = request.POST.get("mois")
        
    montant_ug_im_alls = []
    
    nbug=0
    line_Total_ug=[]
    line_Total_ug.append("TOTAUX")
    for ug in ugs:
        line_Total_ug.append(0)
        nbug = nbug+1
    line_Total_ug.append(0)
        
    for impot in impots:
        line = []
        line.append(impot.impot)
        ids = Impot_Declare.objects.filter(impot = impot)
        for ug in ugs:
            montant = 0
            total_impot = 0 
            for contrib in ug.contribuable_set.all():
                if today.month == 1:
                    declaration = Declaration.objects.filter(contribuable=contrib) &  Declaration.objects.filter(date_limite__month = 12) &  Declaration.objects.filter(date_limite__year = today.year-1)
                    annee = today.year - 1
                else:
                    declaration = Declaration.objects.filter(contribuable=contrib) &  Declaration.objects.filter(date_limite__month = today.month-1) &  Declaration.objects.filter(date_limite__year = today.year)
                    annee = today.year
                m=0
                for id1 in ids:
                    for dec in declaration:
                        if id1.declaration == dec:
                            m = id1.montant
                            break
                         
                montant = montant + m
            line_Total_ug[ug.ug]=line_Total_ug[ug.ug]+montant
            total_impot = total_impot + montant    
            line.append(montant)
        line.append(total_impot)
        line_Total_ug[nbug+1]=line_Total_ug[nbug+1]+total_impot
        montant_ug_im_alls.append(line)
        
    montant_ug_im_alls.append(line_Total_ug)            
    montant_ss_alls = []
    for sous_secteur in sous_secteurs:
        line = []
        line.append(sous_secteur.libelle)
        ids = Impot_Declare.objects.all()
        montant = 0 
        for contrib in sous_secteur.contribuable_set.all():
            if today.month == 1:
                declaration = Declaration.objects.filter(contribuable=contrib) &  Declaration.objects.filter(date_limite__month = 12) &  Declaration.objects.filter(date_limite__year = today.year-1)
            else:
                declaration = Declaration.objects.filter(contribuable=contrib) &  Declaration.objects.filter(date_limite__month = today.month-1) &  Declaration.objects.filter(date_limite__year = today.year)
            m=0
            for id1 in ids:
                for dec in declaration:
                    if id1.declaration == dec:
                        montant = montant + id1.montant
            
        line.append(montant)
        montant_ss_alls.append(line)
    
       
    context = {
        "ugs" : ugs,
        "annee" : annee,
        "mois" : mois[today.month-1],
        "impots" : impots,
        "montant_ug_im_alls" : montant_ug_im_alls,
        "montant_ss_alls" : montant_ss_alls,
        
    }
    return render(request, "soft_cime/stats_perf_gestion.html", context)



# Excel stats 2
@login_required
def excelStats2(request):
    wbl=load_workbook("soft_cime/static/doc/statistiques.xlsx")
    wb= Workbook()
    ws = wbl["Evaluation"]
    
    today = date.today()
    
    def set_border(ws, cell_range):
        border = styles.Border(left=styles.Side(border_style='thin', color='000000'),
                    right=styles.Side(border_style='thin', color='000000'),
                    top=styles.Side(border_style='thin', color='000000'),
                    bottom=styles.Side(border_style='thin', color='000000'))

        rows = ws.iter_rows(cell_range)
        for row in rows:
            for cell in row:
                cell.border = border
            
            
    subtitle = styles.NamedStyle(name = 'subtitle')
    subtitle.font = styles.Font(name = 'Times New Roman', size = 10, bold=True)
    subtitle.fill = styles.PatternFill(patternType = 'solid', fgColor = 'cccccc')

    ws["A20"] = "EVALUATION DES PERFORMANCES DU MOIS DE  "  + str(today.year)
    ws.merge_cells("A20:I20")
    
    ws["A22"] = "I - CELLULE DE GESTION ET DE SUIVI"
    
    ws["A24"] = "RENDEMENT PAR TYPE D'IMPOT, PAR UNITE DE GESTION ET PAR SOUS SECTEUR D'ACTIVITE (VS " + str(today.year)
    
    i=26
    
    ws["A"+str(i)]="UG"
    ws["A"+str(i+1)]="IMPOT"
    ws.column_dimensions['A'].width = 15
    j='B'
    
    
    ugs = UG.objects.all()
    impots = Impot.objects.all()
    dico = {}
    
    
    for ug in ugs:
        ws.merge_cells(j+""+str(i)+":"+j+""+str(i+1))
        ws[j+""+str(i)] = "UG" + str(ug.ug)
        ws[j+""+str(i)].style = subtitle
        ws.column_dimensions[j].width = 15
        dico[ug.ug]=j
        j=chr(ord(j) + 1) 
    
    ws[j+""+str(i)] = "TOTAUX"
    dico["T_impot"]=j
    
    i=i+2
    j='A'
        
    
    for impot in impots:
        ws[j+""+str(i)] = str(impot.impot)
        ws[j+""+str(i)].style=subtitle
        dico[impot.impot]=i
        i=i+1
        
    ws[j+""+str(i)] = "TOTAUX"
    dico["T_ug"]=i    
    
    
    #bordure plage 
    # plage = "A4:"+str(dico["T_impot"])+str(dico["T_ug"])
    # set_border(ws, plage)
    
    # impots_declare = Impot_Declare.objects.filter(declaration in (Declaration.objects.filter(date_limite__month = today.month-1) &  Declaration.objects.filter(date_limite__year = today.year)))
    impots_declare = Impot_Declare.objects.all()

    for impot_dec in impots_declare:
        ref = str(dico[impot_dec.declaration.contribuable.ug.ug]) + str(dico[impot_dec.impot.impot])
        
        if ws[ref].value == None:
            ws[ref]=0
            
        ws[ref]=ws[ref].value + impot_dec.montant
        
        # ws[dico["T_impot"]+""+dico[impot_dec.impot.impot]]="=SUM("
    
    
    wbl.save(os.path.expanduser("~/Downloads/stats_performance.xlsx"))
    os.popen(os.path.expanduser("~/Downloads/stats_performance.xlsx"))
    
    return redirect('stats2')

@login_required
def stat_perf_recette(request):
    if today.month == 1:
        impots_amr = Impot_AMR.objects.filter(Q(date__month = 12) & Q(date__year = today.year-1))
        amrs = AMR.objects.filter(Q(date__month = 12) & Q(date__year = today.year-1)).exclude(impots__count=0)
    else:
        impots_amr = Impot_AMR.objects.filter(Q(date__month = today.month) & Q(date__year = today.year))
        amrs = AMR.objects.filter(Q(date__month = today.month) & Q(date__year = today.year)).exclude()

    stat_recette=[]
    for amr in amrs:
        line=[]
        line.append(amr.recette)
        line.append(amr.contribuable.raison_social + " / " + str(amr.num_amr))
        montant = 0
        montant_budg = 0
        for impot in impots_amr:
            if impot.amr == amr:
                montant = montant + impot.montant
                montant_budg = montant_budg + impot.montant_budg
            
        line.append(montant)
        line.append(montant_budg)
        stat_recette.append(line)
                

        
    montantT = Impot_AMR.objects.aggregate(Sum('montant'))
    montantBT = Impot_AMR.objects.aggregate(Sum('montant_budg'))
        
        
        
            
    context={
        "stat_recette" : stat_recette,
        "montantT" : montantT,
        "montantBT" : montantBT,
        "amrs" : amrs,
    }
    return render(request, "soft_cime/stats_perf_recette.html", context)


def zeroifnone(x):
    if x is None:
        return 0
    else:
        return x
    
@login_required    
def stats_consolide_irc(request, m, y):
    
    wb=load_workbook("soft_cime/static/doc/Statistiques Consolidées 1.xlsx")
    ws = wb.active
    
    mois_stats = m
    an_stats = y

    # if today.month > 1 : 
    #     mois_stats = today.month - 1
    # else:
    #     mois_stats = 12
        
    payements = Payement.objects.filter(Q(date__month = mois_stats) and Q(date__year = an_stats))
    declaration = []
    for payement in payements:
        dec = Declaration.objects.filter(num_avis = payement.num_avis)
        for d in dec:
            declaration.append(d)
            
    amrs = AMR.objects.filter(date__month=mois_stats)
    ws['G9'] = mois[mois_stats]
    ws['C9'] = an_stats
    j='B'
    i=16
    
    ws.oddFooter.center.text ="@Le Gestionnaire \n" + str(datetime.today())
    
    #Traitements et salaires
    ws[j+str(i)] = Impot.objects.get(Q(impot='IRPP') and Q(impot="SOLDE IRPP")).impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(Q(impot='IRPP') and Q(impot="SOLDE IRPP")).impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    #Revenus Foncier
    j=chr(ord(j) - 1)
    i=i+5
    ws[j+str(i)] = Impot.objects.get(impot='RF').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='RF').impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    #Revenus Capitaux Mobilier
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = Impot.objects.get(impot='IRCM').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='IRCM').impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    #Accompte Mensuel BIC
    j=chr(ord(j) - 1)
    i=i+2
    ws[j+str(i)] = Impot.objects.get(Q(impot='AIR BIC') and Q(impot="SOLDE BIC")).impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(Q(impot='AIR BIC') and Q(impot="SOLDE BIC")).impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    #Accompte Mensuel BNC
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = Impot.objects.get(Q(impot='AIR BNC') and Q(impot="SOLDE BNC")).impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(Q(impot='AIR BNC') and Q(impot="SOLDE BNC")).impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    
    #TVA
    j=chr(ord(j) - 1)
    i=i+7
    ws[j+str(i)] = Impot.objects.get(Q(impot='TVA') and Q(impot="SOLDE TVA")).impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(Q(impot='TVA') and Q(impot="SOLDE TVA")).impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    #Droits d'assise
    j=chr(ord(j) - 1)
    i=i+3
    ws[j+str(i)] = Impot.objects.get(impot='ASSISE').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='ASSISE').impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    #Taxe de Séjour
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = Impot.objects.get(impot='TAXE DE SEJOUR').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='TAXE DE SEJOUR').impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    # Solde Impot sur les societés
    j=chr(ord(j) - 1)
    i=i+2
    ws[j+str(i)] = Impot.objects.get(impot='SOLDE IS/IRCM').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='SOLDE IS/IRCM').impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    #Accompte Mensuel
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = Impot.objects.get(impot='IS').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='IS').impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    #Accompte Loyer
    j=chr(ord(j) - 1)
    i=i+5
    ws[j+str(i)] = Impot.objects.get(impot='PSI').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='PSI').impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    #Précompte achats
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = Impot.objects.get(Q(impot='PSA') and Q(impot="SOLDE PSA")).impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(Q(impot='PSA') and Q(impot="SOLDE PSA")).impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    
    #CAC TVA FEICOM
    j=chr(ord(j) - 1)
    i=i+8
    ws[j+str(i)] = zeroifnone(Impot.objects.get(Q(impot='TVA') and Q(impot="SOLDE TVA")).impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC FEI').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(Q(impot='TVA') and Q(impot="SOLDE TVA")).impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC FEI').taux
    
    
    #CAC TVA COMMUNE
    j=chr(ord(j) - 1)
    i=i+2
    ws[j+str(i)] = zeroifnone(Impot.objects.get(Q(impot='TVA') and Q(impot="SOLDE TVA")).impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC COM').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(Q(impot='TVA') and Q(impot="SOLDE TVA")).impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC COM').taux
    
    
    #CAC IRPP FEICOM
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = zeroifnone(Impot.objects.get(Q(impot='IRPP') and Q(impot="SOLDE IRPP")).impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC FEI').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(Q(impot='IRPP') and Q(impot="SOLDE IRPP")).impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC FEI').taux
    
    
    #CAC IRPP COMMUNE
    j=chr(ord(j) - 1)
    i=i+2
    ws[j+str(i)] = zeroifnone(Impot.objects.get(Q(impot='IRPP') and Q(impot="SOLDE IRPP")).impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC COM').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(Q(impot='IRPP') and Q(impot="SOLDE IRPP")).impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC COM').taux
            
    
    #CAC IS FEICOM
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = zeroifnone(Impot.objects.get(Q(impot='IS') and Q(impot="SOLDE IS/IRCM")).impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC FEI').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(Q(impot='IS') and Q(impot="SOLDE IS/IRCM")).impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC FEI').taux
    
    
    #CAC IS COMMUNE
    j=chr(ord(j) - 1)
    i=i+2
    ws[j+str(i)] = zeroifnone(Impot.objects.get(Q(impot='IS') and Q(impot="SOLDE IS/IRCM")).impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC COM').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(Q(impot='IS') and Q(impot="SOLDE IS/IRCM")).impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC COM').taux
    
    #CRTV
    j=chr(ord(j) - 1)
    i=i+5
    ws[j+str(i)] = Impot.objects.get(impot='RAV').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='RAV').impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    #CCF
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = Impot.objects.get(impot='CCF').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='CCF').impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    #FNE
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = Impot.objects.get(impot='FNE').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='FNE').impot_amr_set.filter(Q(date__month = mois_stats) and Q(date__year = an_stats)).aggregate(Sum('montant'))['montant__sum']
    
    
    wb.save(os.path.expanduser("~/Downloads/Statistiques Consolidée Impôts sur le revenu - " + mois[mois_stats] + ".xlsx"))
    os.popen(os.path.expanduser("~/Downloads/Statistiques Consolidée Impôts sur le revenu - " + mois[mois_stats] + ".xlsx"))
    
    return redirect('stats_etats')




@login_required
def stats_consolide_retc(request, m, y):
    
    wb=load_workbook("soft_cime/static/doc/Statistiques Consolidées 2.xlsx")
    ws = wb.active
    
    mois_stats = m
    an_stats = y
    
    # if today.month > 1 : 
    #     mois_stats = today.month - 1
    # else:
    #     mois_stats = 12
        
    payements = Payement.objects.filter(date__month = mois_stats).filter(date__year = an_stats)
    declaration = []
    for payement in payements:
        dec = Declaration.objects.filter(num_avis = payement.num_avis).filter(contribuable=payement.contribuable)
        for d in dec:
            declaration.append(d)
            
    amrs = AMR.objects.filter(date__month=mois_stats).filter(date__year = an_stats)
    
    ws['G9'] = mois[mois_stats]
    ws['C9'] = an_stats
    
    j='B'
    i=20
    
    ws.oddFooter.center.text ="@Le Gestionnaire \n" + str(datetime.today())
    
    #Autres mutations
    ws[j+str(i)] = Impot.objects.get(impot='AUTRE DF').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='AUTRE DF').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']
    
    #Taxe à l'Essieu
    j=chr(ord(j) - 1)
    i=i+2
    ws[j+str(i)] = Impot.objects.get(impot='TAE').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='TAE').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']
    
    #Timbres connaissements et contrat transport
    j=chr(ord(j) - 1)
    i=i+8
    ws[j+str(i)] = Impot.objects.get(impot='TIM CT').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='TIM CT').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']
    
    #TAXE/JEUX
    j=chr(ord(j) - 1)
    i=i+3
    ws[j+str(i)] = Impot.objects.get(impot='TAXE/JEUX').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='TAXE/JEUX').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']
    
    #Redevance de superficie (RFA)
    j=chr(ord(j) - 1)
    i=i+2
    ws[j+str(i)] = Impot.objects.get(impot='RFA').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='RFA').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']
    
    
    #Taxe d'abbatage (TA)
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = Impot.objects.get(impot='TA').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='TA').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']
    
    #Redevance de superficie annuelle (RSA)
    j=chr(ord(j) - 1)
    i=i+4
    ws[j+str(i)] = Impot.objects.get(impot='RSA').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='RSA').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']
    
    #Droit fixe (DF)
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = Impot.objects.get(impot='DF').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='DF').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']
    
    # Taxe advalorem sur les recettes minères (TAV)
    j=chr(ord(j) - 1)
    i=i+3
    ws[j+str(i)] = Impot.objects.get(impot='TAV').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='TAV').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']
    
    #Taxe d'insp. Sani. Vétér. Commerce (TISV)
    j=chr(ord(j) - 1)
    i=i+4
    ws[j+str(i)] = Impot.objects.get(impot='TISV').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='TISV').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']
    
    #Droits d'exploitation de la pêche (TE)
    j=chr(ord(j) - 1)
    i=i+2
    ws[j+str(i)] = Impot.objects.get(impot='TE').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='TE').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']
    
    #Timbre Automobile (Vignette)
    j=chr(ord(j) - 1)
    i=i+4
    ws[j+str(i)] = Impot.objects.get(impot='VIGNETTE').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']
    j=chr(ord(j) + 1)
    ws[j+str(i)] = Impot.objects.get(impot='VIGNETTE').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']
    
    
    #BAUX FEICOM
    j=chr(ord(j) - 1)
    i=i+4
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='BAIL').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='FEI 18%').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='BAIL').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='FEI 18%').taux
    
    
    #BAUX CU
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='BAIL').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CU 18%').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='BAIL').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CU 18%').taux
    
    
    
    #BAUX COMMUNE
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='BAIL').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='COM 54%').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='BAIL').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='COM 54%').taux
    
    
    #TPF FEICOM
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='TPF').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='FEI 18%').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='TPF').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='FEI 18%').taux
    
    
    #TPF CU
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='TPF').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CU 18%').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='TPF').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CU 18%').taux
    
    
    
    #TPF COMMUNE
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='TPF').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='COM 54%').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='TPF').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='COM 54%').taux
            
    
    #RFA FEICOM
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='RFA').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='RFA FEI').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='RFA').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='RFA FEI').taux
    
    
    #RFA COMMUNE
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='RFA').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='RFA COM').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='RFA').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='RFA COM').taux
    
    
    #CAC TISV
    j=chr(ord(j) - 1)
    i=i+2
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='TISV').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='TISV').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC').taux
    
    #CAC TE
    j=chr(ord(j) - 1)
    i=i+1
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='TE').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC').taux
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='TE').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']) * Part_Impot.objects.get(nom='CAC').taux
    
    
    #TAV COM
    j=chr(ord(j) - 1)
    i=i+2
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='TAV COM').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum'])
    j=chr(ord(j) + 1)
    ws[j+str(i)] = zeroifnone(Impot.objects.get(impot='TAV COM').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])
    
    #MONTANT IMPOTS BUDGETAIRE ET NON BUDGETAIRE
    i=i+8
    ws[j+str(i)] = zeroifnone(Impot_Declare.objects.filter(Q(declaration__in = declaration) and Q(impot__type_impot='Budgétaire')).aggregate(Sum('montant'))['montant__sum'])
    i=i+1
    ws[j+str(i)] = zeroifnone(Impot_Declare.objects.filter(Q(declaration__in = declaration) and Q(impot__type_impot='Non Budgétaire')).aggregate(Sum('montant'))['montant__sum'])
    
    
    
    wb.save(os.path.expanduser("~/Downloads/Statistiques Consolidée Recettes d'enregistrement - " + mois[mois_stats] + " " + an_stats +  ".xlsx"))
    os.popen(os.path.expanduser("~/Downloads/Statistiques Consolidée Recettes d'enregistrement - " + mois[mois_stats] + " " + an_stats +  ".xlsx"))
    
    return redirect('stats_etats')



@login_required
def stats_recette_af(request, m, y):
    
    wb=load_workbook("soft_cime/static/doc/R. AFFECTEES.xlsx")
    ws = wb.active
    
    mois_stats = m
    an_stats = y
    
        
    payements = Payement.objects.filter(date__month = mois_stats).filter(date__year = an_stats)
    declaration = []
    for payement in payements:
        dec = Declaration.objects.filter(num_avis = payement.num_avis).filter(contribuable=payement.contribuable)
        for d in dec:
            declaration.append(d)
            
    amrs = AMR.objects.filter(date__month=mois_stats).filter(date__year = an_stats)
    
    ws['C11'] = "EXERCICE : " + str(an_stats)
    ws['E11'] = "MOIS : " + mois[mois_stats]
    
    j='D'
    i=20
    
    ws.oddFooter.center.text ="@Le Gestionnaire \n" + str(datetime.today())
    
    #RFA FRAIS D'ASSIETTE
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='RFA').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='RFA').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='RFA FAR').taux
    #RFA FEICOM
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='RFA').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='RFA').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='RFA FEI').taux
    #RFA COMMUNE
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='RFA').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='RFA').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='RFA COM').taux
    #RFA ETAT/BUDGET
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='RFA').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='RFA').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='RFA Principale').taux
    
    
    #TVA FRAIS D'ASSIETTE
    i=i+2
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='TVA').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='TVA').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='FAR').taux
    #TVA FEICOM
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='TVA').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='TVA').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='CAC FEI').taux
    #TVA COMMUNE
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='TVA').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='TVA').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='CAC COM').taux
    
    
    #IRPP FRAIS D'ASSIETTE
    i=i+2
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='IRPP').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='IRPP').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='FAR').taux
    #IRPP FEICOM
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='IRPP').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='IRPP').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='CAC FEI').taux
    #IRPP COMMUNE
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='IRPP').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='IRPP').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='CAC COM').taux
    
    
    #IS FRAIS D'ASSIETTE
    i=i+2
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='IS').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='IS').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='FAR').taux
    #IS FEICOM
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='IS').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='IS').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='CAC FEI').taux
    #IS COMMUNE
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='IS').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='IS').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='CAC COM').taux
    
    
    # PATENTE 
    
    
    
    
    
    # LICENCE
    
    
    
    
    #TPF FRAIS D'ASSIETTE
    i=i+18
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='TPF').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='TPF').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='FAR 10%').taux
    #TPF FEICOM
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='TPF').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='TPF').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='FEI 18%').taux
    #TPF CU
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='TPF').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='TPF').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='CU 18%').taux
    #TPF COM
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='TPF').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='TPF').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='COM 54%').taux
    
    
    #BAIL FRAIS D'ASSIETTE
    i=i+2
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='BAIL').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='BAIL').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='FAR 10%').taux
    #BAIL FEICOM
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='BAIL').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='BAIL').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='FEI 18%').taux
    #BAIL CU
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='BAIL').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='BAIL').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='CU 18%').taux
    #BAIL COM
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='BAIL').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='BAIL').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='COM 54%').taux
    
    
    #TAV FRAIS D'ASSIETTE
    i=i+21
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='TAV').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='TAV').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='TAV IMPOT').taux
    #TAV COMMUNE
    i=i+21
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='TAV').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='TAV').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum'])) * Part_Impot.objects.get(nom='TAV COM').taux
    #TAV BUDGET
    i=i+21
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='TAV').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='TAV').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']))
    
    #CFC
    i=i+6
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='CFC').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='CFC').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']))
    #FNE
    i=i+1
    ws[j+str(i)] = (zeroifnone(Impot.objects.get(impot='FNE').impot_declare_set.filter(declaration__in = declaration).aggregate(Sum('montant'))['montant__sum']) + zeroifnone(Impot.objects.get(impot='FNE').impot_amr_set.filter(amr__in = amrs).aggregate(Sum('montant'))['montant__sum']))
    
    
    wb.save(os.path.expanduser("~/Downloads/Recette Affectées - " + mois[mois_stats] + " " + str(an_stats) + ".xlsx"))
    os.popen(os.path.expanduser("~/Downloads/Recette Affectées - " + mois[mois_stats] + " " + str(an_stats) +  ".xlsx"))
    
    return redirect('stats_etats')



@login_required
def excel_virements(request, m, y):
    payements = Payement.objects.filter(date_virement__year = y).filter(date_virement__month = m).order_by("date_virement")
    wb=load_workbook("soft_cime/static/doc/Etats des virements.xlsx")
    ws = wb.active
    
    
    borderStyle = styles.Side(style = 'thin', color = '000000')
    s = styles.NamedStyle(name = 's')
    # s.font = styles.Font(name = 'Times New Roman', size = 10, bold=True)
    s.border = styles.Border(left = borderStyle, right = borderStyle, top = borderStyle, bottom = borderStyle)
    # s.alignment.horizontal = "center"
    
    
    
    ws["F15"] = today
    ws["D18"] = "01/" + str(m) + "/" + str(y) + " AU "
    
    if m==today.month and y==today:
        ws["E18"] = str(today.day) + "/" + str(m) + "/" + str(y)
    else:
        ws["E18"] = "15/" + str(m) + "/" + str(y)
    
    j='A'
    i=21
    
    montant_compta = 0
    montant_non_compta = 0
    
    compt = 1
    for payement in payements:
        ws[j+str(i)] = compt
        ws[j+str(i)].style = s
        j=chr(ord(j)+1)
        
        ws[j+str(i)] = payement.contribuable.raison_social
        ws[j+str(i)].style = s
        j=chr(ord(j)+1)
        
        ws[j+str(i)] = str(payement.date_virement)
        ws[j+str(i)].style = s
        j=chr(ord(j)+1)
        
        ws[j+str(i)] = payement.num_virement
        ws[j+str(i)].style = s
        j=chr(ord(j)+1)
        
        ws[j+str(i)] = payement.montant
        ws[j+str(i)].style = s
        j=chr(ord(j)+1)
        
        ws[j+str(i)] = payement.etat
        ws[j+str(i)].style = s
        
        if payement.etat == "Comptabilisé":
            montant_compta = montant_compta + payement.montant
        else:
            montant_non_compta = montant_compta + payement.montant
            
        j=chr(ord(j)-5)
        i=i+1
        compt = compt + 1
    
    i=i+2
    j=chr(ord(j)+2)
    ws[j+str(i)] = "Comptabilisé"
    ws[j+str(i)].style = s
    j=chr(ord(j)+1)
    ws[j+str(i)] = montant_compta
    ws[j+str(i)].style = s
    
    i=i+1
    j=chr(ord(j)-1)
    ws[j+str(i)] = "Non Comptabilisé"
    ws[j+str(i)].style = s
    j=chr(ord(j)+1)
    ws[j+str(i)] = montant_non_compta
    ws[j+str(i)].style = s
    
    i=i+1
    j=chr(ord(j)-1)
    ws[j+str(i)] = "Total"
    ws[j+str(i)].style = s
    j=chr(ord(j)+1)
    ws[j+str(i)] = montant_non_compta + montant_compta
    ws[j+str(i)].style = s
    
        
    wb.save(os.path.expanduser("~/Downloads/Etats des virements - " + mois[m] + " " + str(y) + ".xlsx"))
    os.popen(os.path.expanduser("~/Downloads/Etats des virements - " + mois[m] + " " +  str(y) +".xlsx"))
    
    return redirect('stats_etats')
    
    

@login_required
def stats_etats(request):
    
    if request.method == "POST":
        num_m = int(request.POST.get("date"))
        an = int(request.POST.get("an"))
    else : 
        if today.month > 1:
            num_m = int(today.month)-1
            an = today.year
        else:
            num_m = 12
            an = today.year - 1
        
    range_an = []
    for i in range(today.year, today.year - 10, -1):
        range_an.append(i)
            
    context = {
        "mois" : mois[num_m],
        "num_m" : num_m,
        "today" : today,
        "range_an": range_an,
        "an" : an,
    }
    return render(request, "soft_cime/stats_etats.html", context)


# AMR
@login_required
def new_amr(request):
    
    contribuables = contribuable.objects.all()
    message = ""
    success = False
    if request.method == 'POST':
        num_amr = request.POST.get("num_amr")
        contrib = contribuable.objects.get(NIU=request.POST.get("contribuable"))
        personnel = Personnel.objects.get(user = request.user)
        recette = request.POST.get("recette")
        newAMR = AMR.objects.create(num_amr=num_amr,contribuable=contrib, personnel=personnel, recette=recette)        
        newAMR.save()
        message = 'Success'
        success = True
        return redirect('new_impot_amr',newAMR.id)
        
    amrs = AMR.objects.all()
    nb_amr_in = 0
    for amr in amrs:
        if amr.impots.count() == 0:
            nb_amr_in = nb_amr_in + 1
            
    context = {
        'contribuables' : contribuables,
        'message' : message,
        'success' : success,
        'nb_amr_in' : nb_amr_in,
    }
    return render(request, "soft_cime/amr-new.html", context)


@login_required
def new_impot_amr(request, idAmr):
    amr = AMR.objects.get(id=idAmr)
    impots = Impot.objects.all()
    impots_amr = Impot_AMR.objects.filter(amr=amr)
    print(impots_amr)
    message = ""
    classmsg = ""
    if request.method == 'POST':
        montant = request.POST.get("montant")
        montant_budg = request.POST.get("recette_budg")
        impot = Impot.objects.get(impot=request.POST.get("impot"))
        
        exist = Impot_AMR.objects.filter(impot=impot) & Impot_AMR.objects.filter(amr=amr)
        
        if len(exist)==0:
            newImpotAMR = Impot_AMR.objects.create(amr=amr,impot=impot, montant=montant,montant_budg = montant_budg)        
            newImpotAMR.save()
            message = 'Success'
            classmsg = "text-success"
        else:
            message = 'Vous essayer de dupliquer une déclaration !!!'
            classmsg = "text-danger"
            
    amrs = AMR.objects.all()
    nb_amr_in = 0
    for amr in amrs:
        if amr.impots.count() == 0:
            nb_amr_in = nb_amr_in + 1
            
    context = {
        'impots' : impots,
        'message' : message,
        'classmsg' : classmsg,
        'nb_amr_in' : nb_amr_in,
        'impots_amr' : impots_amr,
    }
    return render(request, "soft_cime/amr-new1.html", context)

@login_required
def liste_amr(request):
    
    amrs = AMR.objects.all().order_by('-date')
    nbamr = AMR.objects.count()
    impots_amr = Impot_AMR.objects.all()

    amrs = AMR.objects.all()
    nb_amr_in = 0
    for amr in amrs:
        if amr.impots.count() == 0:
            nb_amr_in = nb_amr_in + 1
            
    context = {
        'amrs' : amrs,
        'nbamr' : nbamr,
        'impots_amr' : impots_amr,
        'nb_amr_in' : nb_amr_in,
    }
    
        
    return render(request, "soft_cime/amr-history.html", context)

@login_required
def amr_incomplet(request):
    amr_history = AMR.objects.filter().order_by('-date')
    impots_amr = Impot_AMR.objects.all()
    amrs = AMR.objects.all()
    nb_amr_in = 0
    for amr in amrs:
        if amr.impots.count() == 0:
            nb_amr_in = nb_amr_in + 1


    context = {
        'amrs' : amr_history,
        'impots_amr' : impots_amr,
        'nb_amr_in' : nb_amr_in,
    }
    
    
    return render(request, "soft_cime/amr-incomplet.html", context)


@login_required
def update_amr(request, idAmr):
    contribuables = contribuable.objects.all()
    amr = AMR.objects.get(id=idAmr)
    message = ""
    classmsg=""
    if request.method == 'POST':
        amr.num_amr = request.POST.get("num_amr")
        amr.contribuable = contribuable.objects.get(NIU=request.POST.get("contribuable"))
        amr.recette = request.POST.get("recette")
        amr.save()
        message = 'Success'
        classmsg="text-success"
        return redirect('new_impot_amr',amr.id)
    
    amrs = AMR.objects.all()
    nb_amr_in = 0
    for amr in amrs:
        if amr.impots.count() == 0:
            nb_amr_in = nb_amr_in + 1
        
    
    context = {
        'contribuables' : contribuables,
        'amr':amr,
        'message' : message,
        'classmsg' : classmsg,
        'nb_amr_in' : nb_amr_in,
    }
    return render(request, "soft_cime/amr-update.html", context)

@login_required
def detail_amr(request, idAmr):
    amr = AMR.objects.get(id=idAmr)
    impots_amr = Impot_AMR.objects.filter(amr=amr)
    
    amrs = AMR.objects.all()
    nb_amr_in = 0
    for amr in amrs:
        if amr.impots.count() == 0:
            nb_amr_in = nb_amr_in + 1
            
            
    context = {
        'amr' : amr,
        'impots_amr' : impots_amr,
        'nb_amr_in' : nb_amr_in,
    }
    
    

    return render(request, "soft_cime/amr-detail.html", context)


@login_required
def excel_amr(request,idAmr):
    
    amr = AMR.objects.get(id=idAmr)
    impots_amr = Impot_AMR.objects.filter(amr=amr)
    context = {
        'amr' : amr,
        'impots_amr' : impots_amr,
    }
    wb=load_workbook("soft_cime/static/doc/detail amr.xlsx")
    ws = wb.active
    
    ws['B3'] = str(amr.num_amr)
    ws['B4'] = str(amr.contribuable.NIU) + " : " + str(amr.contribuable.raison_social)
    ws['B5'] = str(amr.date)
    ws['B6'] = amr.recette
    ws['B7'] = amr.personnel.nom
    
    
    j='A'
    i=11
    for impot in impots_amr:
        ws["A"+str(i)]=impot.impot.impot
        ws["B"+str(i)]=impot.montant
        ws["C"+str(i)]=impot.montant_budg
        i = i+1
        # j=chr(ord(j) + 1)

    ws.oddFooter.center.text ="Powered by Le Gestionnaire \n" + str(datetime.today())
    wb.save(os.path.expanduser("~/Downloads/AMR " + str(amr.num_amr) + ".xlsx"))
    os.popen(os.path.expanduser("~/Downloads/AMR " + str(amr.num_amr) + ".xlsx"))
    
    return redirect("detail_amr", idAmr)


    
    
    